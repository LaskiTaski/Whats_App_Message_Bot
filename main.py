import os
import random

from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from random import randint
import openpyxl
import json


options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
# options.add_argument('user-data-dir=C:\\Users\\Tim\\AppData\\Local\\Google\\Chrome\\User Data')


workbook = openpyxl.load_workbook("number_list.xlsx")
worksheet = workbook.active


def load_numbers(): #Считывание всех номеров из xlsx
    list_number = [col[i].value for i in range(1, worksheet.max_row) for col in worksheet.iter_cols(1, worksheet.max_column) if col[i].value]
    return list_number

def click_new_chat():# Нажатие на кнопку новый чат
    browser.find_element(By.XPATH, '//*[@aria-label="Новый чат"]').click()
    sleep(randint(2, 8))

def entering_number(number):# Вставляем номер телефона
    browser.find_element(By.XPATH, '//*[@title="Текстовое поле поиска"]').send_keys(number)
    sleep(randint(2, 8))

def click_user_account():# Выбираем аккаунт "кому отправить"
    browser.find_element(By.XPATH, '//*[@tabindex="0" and @class="_ak72 _ak73" and @role="button"]').click()
    sleep(randint(2, 8))

def entering_message(message): # Пишем текст
    browser.find_element(By.XPATH, '//*[@title="Введите сообщение"]').send_keys(message)
    sleep(randint(2, 8))

def click_send_message(): # Нажатие на кнопку отправить
    browser.find_element(By.XPATH, '//*[@aria-label="Отправить"]').click()
    sleep(randint(2, 8))

def click_back(): # Вернуться назад если пользователь не найден
    browser.find_element(By.XPATH, '//*[@aria-label="Назад"]').click()
    sleep(2)

result_json = {}
all_message = [
    'Здравствуйте, меня зовут Артём, я ваш куратор по программе Код Будущего Maximum Education, вам нужно посетить личный кабинет и перейти по ссылке, чтобы присоединиться к группе.',
    'Приветствую вас! Я Артём, ваш куратор в программе Код Будущего от Maximum Education. Прошу зайти в личный кабинет и кликнуть по ссылке, чтобы стать членом нашей группы',
    'Здравствуйте, уважаемый участник. Я Артём, ваш куратор в проекте "Код Будущего" от Maximum Education. Вам необходимо авторизоваться в личном кабинете и пройти по предложенной ссылке, чтобы присоединиться к нашей учебной группе',
    'Привет, Артём с тобой! Я являюсь твоим персональным куратором в рамках программы "Код Будущего" Maximum Education. Не забудь войти в свой личный кабинет и перейти по ссылке для зачисления в группу обучения',
    'Доброго времени суток! Я Артём, ваш куратор в образовательной программе "Код Будущего" от Maximum Education. Пожалуйста, посетите свой личный кабинет и кликните на предоставленную ссылку, дабы стать полноценным участником нашей группы',
    'Добрый день! я Артём, ваш куратор в проекте "Код Будущего", реализуемом Maximum Education. Для начала работы в нашей программе, вам требуется авторизоваться в личном кабинете и перейти по предложенному адресу ссылки, чтобы стать частью учебной команды.'
]

if __name__ == '__main__':
    list_number = load_numbers()
    with webdriver.Chrome(options=options) as browser:
        browser.get('https://web.whatsapp.com/')
        sleep(30)
        for number in list_number:
            click_new_chat()
            entering_number(number)

            try:
                message = random.choice(all_message)
                click_user_account()
                entering_message(message)
                click_send_message()
                print(f'Пользователь {number} Получил сообщение\n')
                result_json[number] = True

            except:
                click_back()
                print(f'Пользователь {number} Не найден\n')
                result_json[number] = '-------------------FALSE-------------------------'

            finally:
                with open('res.json', 'a', encoding='utf-8') as file:
                    json.dump(result_json, file, indent=4, ensure_ascii=False)
                    result_json = {}
